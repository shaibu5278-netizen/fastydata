from flask import Flask, render_template, request, flash, redirect, url_for, abort, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from forms.registerform import RegisterationForm, EditUserForm
from forms.loginform import LoginForm
from models.user import db, User
from models.account import Account, Package, PaymentType, Transaction
from flask_migrate import Migrate
import os
from forms.transactionform import TransactionForm
from forms.packageform import PackageForm, PackageItemForm
from forms.paymenttypeform import PaymentTypeForm
import csv
from io import StringIO
from flask import Response
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate
from models.account import PackageItem
import uuid
from sqlalchemy.exc import IntegrityError
from flask import jsonify
from flask_login import login_required, current_user
from models.account import db, Account, Package, PackageItem, Orders
from functools import wraps
from datetime import datetime
from flask import send_file
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from sqlalchemy import desc, asc
import re

app = Flask(__name__)
app.config['SECRET_KEY'] = 'MY_KEY'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db.init_app(app)
migrate=Migrate(app,db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# # Create database tables
# with app.app_context():
#     db.create_all()

@app.route("/")
def home():
    packages = Package.query.filter_by(is_active=True).all()
    return render_template("home.html", packages=packages)

@app.route("/register", methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = RegisterationForm()
    if form.validate_on_submit():
        user = User(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            username=form.username.data,
            email=form.email.data,
            phone_number=form.phone_number.data,
            user_type='user'  # Default user type for new registrations
        )
        user.set_password(form.password.data)
        db.session.add(user)
        try:
            db.session.commit()
            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('Username or email is already in use. Please choose a different one.', 'danger')
            else:
                flash('An error occurred during registration. Please try again.', 'danger')
    return render_template("register.html", form=form)

@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = LoginForm()
    if form.validate_on_submit():
        user = db.session.query(User).filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            return redirect(url_for('home'))
        flash('Invalid username or password', 'error')
    return render_template("login.html", form=form)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for('home'))

@app.route("/users")
@login_required
def users():
    user_list = User.query.all()
    return render_template("users.html", users=user_list)

@app.route("/users/edit/<int:user_id>", methods=["GET", "POST"])
@login_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = EditUserForm(obj=user)
    if form.validate_on_submit():
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.username = form.username.data
        user.email = form.email.data
        user.phone_number = form.phone_number.data
        user.user_type = form.user_type.data
        if form.password.data:
            user.set_password(form.password.data)
        try:
            db.session.commit()
            flash('User updated successfully!', 'success')
            return redirect(url_for('users'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('Username or email is already in use. Please choose a different one.', 'danger')
            else:
                flash('An error occurred while updating the user. Please try again.', 'danger')
    return render_template("edit_user.html", form=form, user=user)

@app.route("/users/delete/<int:user_id>", methods=["POST"])
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('users'))
    
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.first_name} {user.last_name} has been deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting user. Please try again.', 'error')
    
    return redirect(url_for('users'))

@app.route("/transactions")
@login_required
def transactions():
    # Show all transactions for the current user's account
    account = Account.query.filter_by(user_id=current_user.id).first()
    if not account:
        flash('No account found for this user.', 'error')
        return redirect(url_for('home'))
    txns = Transaction.query.filter_by(account_id=account.id).order_by(Transaction.transaction_date.desc()).all()
    return render_template("transactions.html", transactions=txns)

@app.route("/transactions/add", methods=["GET", "POST"])
@login_required
def add_transaction():
    account = Account.query.filter_by(user_id=current_user.id).first()
    if not account:
        flash('No account found for this user.', 'error')
        return redirect(url_for('home'))
    form = TransactionForm()
    # Populate dropdowns
    form.payment_type.choices = [(pt.id, pt.name) for pt in PaymentType.query.filter_by(is_active=True).all()]
    if form.validate_on_submit():
        # Calculate new balance
        if form.transaction_type.data == 'credit':
            new_balance = account.balance + form.amount.data
        else:
            new_balance = account.balance - form.amount.data
        txn = Transaction(
            account_id=account.id,
            payment_type_id=form.payment_type.data,
            amount=form.amount.data,
            transaction_type=form.transaction_type.data,
            reference_number=str(uuid.uuid4()),
            balance_after=new_balance,
            description=form.description.data,
        )
        account.balance = new_balance
        db.session.add(txn)
        try:
            db.session.commit()
            flash('Transaction added successfully!', 'success')
            return redirect(url_for('transactions'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('Reference number already exists. Please use a unique one.', 'danger')
            else:
                flash('An error occurred while adding the transaction. Please try again.', 'danger')
    return render_template("add_transaction.html", form=form)

# --- Package CRUD ---
@app.route('/packages')
@login_required
def packages():
    q = request.args.get('q', '')
    query = Package.query
    if q:
        query = query.filter(Package.name.ilike(f'%{q}%'))
    packages = query.order_by(Package.name).all()
    return render_template('packages.html', packages=packages, q=q)

@app.route('/packages/add', methods=['GET', 'POST'])
@login_required
def add_package():
    form = PackageForm()
    if form.validate_on_submit():
        image_filename = None
        if form.image.data:
            from werkzeug.utils import secure_filename
            import os
            filename = secure_filename(form.image.data.filename)
            image_path = os.path.join('static', 'package_images', filename)
            form.image.data.save(image_path)
            image_filename = filename
        package = Package(
            name=form.name.data,
            description=form.description.data,
            image_filename=image_filename,
            is_active=form.is_active.data
        )
        db.session.add(package)
        try:
            db.session.commit()
            flash('Package added!', 'success')
            return redirect(url_for('packages'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A package with that name already exists.', 'danger')
            else:
                flash('An error occurred while adding the package. Please try again.', 'danger')
    return render_template('add_edit_package.html', form=form, title='Add Package', package=None)

@app.route('/packages/edit/<int:package_id>', methods=['GET', 'POST'])
@login_required
def edit_package(package_id):
    package = Package.query.get_or_404(package_id)
    form = PackageForm(obj=package)
    if form.validate_on_submit():
        package.name = form.name.data
        package.description = form.description.data
        if form.image.data:
            from werkzeug.utils import secure_filename
            import os
            filename = secure_filename(form.image.data.filename)
            image_path = os.path.join('static', 'package_images', filename)
            form.image.data.save(image_path)
            package.image_filename = filename
        package.is_active = form.is_active.data
        try:
            db.session.commit()
            flash('Package updated!', 'success')
            return redirect(url_for('packages'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A package with that name already exists.', 'danger')
            else:
                flash('An error occurred while updating the package. Please try again.', 'danger')
    return render_template('add_edit_package.html', form=form, title='Edit Package', package=package)

@app.route('/packages/delete/<int:package_id>', methods=['POST'])
@login_required
def delete_package(package_id):
    package = Package.query.get_or_404(package_id)
    db.session.delete(package)
    try:
        db.session.commit()
        flash('Package deleted!', 'success')
    except IntegrityError as e:
        db.session.rollback()
        if 'FOREIGN KEY constraint failed' in str(e.orig):
            flash('Cannot delete package as it is linked to transactions or items.', 'danger')
        else:
            flash('An error occurred while deleting the package. Please try again.', 'danger')
    return redirect(url_for('packages'))

# --- PaymentType CRUD ---
@app.route('/payment_types')
@login_required
def payment_types():
    q = request.args.get('q', '')
    query = PaymentType.query
    if q:
        query = query.filter(PaymentType.name.ilike(f'%{q}%'))
    payment_types = query.order_by(PaymentType.name).all()
    return render_template('payment_types.html', payment_types=payment_types, q=q)

@app.route('/payment_types/add', methods=['GET', 'POST'])
@login_required
def add_payment_type():
    form = PaymentTypeForm()
    if form.validate_on_submit():
        payment_type = PaymentType(
            name=form.name.data,
            description=form.description.data,
            is_active=form.is_active.data
        )
        db.session.add(payment_type)
        try:
            db.session.commit()
            flash('Payment type added successfully!', 'success')
            return redirect(url_for('payment_types'))
        except IntegrityError as e:
            db.session.rollback()
            # You can check for unique constraint here if you want a custom message
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A payment type with that name already exists.', 'danger')
            else:
                flash('An error occurred while adding the payment type.', 'danger')
    return render_template('add_edit_payment_type.html', form=form, title='Add Payment Type')

@app.route('/payment_types/edit/<int:payment_type_id>', methods=['GET', 'POST'])
@login_required
def edit_payment_type(payment_type_id):
    payment_type = PaymentType.query.get_or_404(payment_type_id)
    form = PaymentTypeForm(obj=payment_type)
    if form.validate_on_submit():
        payment_type.name = form.name.data
        payment_type.description = form.description.data
        payment_type.is_active = form.is_active.data
        try:
            db.session.commit()
            flash('Payment type updated!', 'success')
            return redirect(url_for('payment_types'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A payment type with that name already exists.', 'danger')
            else:
                flash('An error occurred while updating the payment type. Please try again.', 'danger')
    return render_template('add_edit_payment_type.html', form=form, title='Edit Payment Type')

@app.route('/payment_types/delete/<int:payment_type_id>', methods=['POST'])
@login_required
def delete_payment_type(payment_type_id):
    payment_type = PaymentType.query.get_or_404(payment_type_id)
    db.session.delete(payment_type)
    try:
        db.session.commit()
        flash('Payment type deleted!', 'success')
    except IntegrityError as e:
        db.session.rollback()
        if 'FOREIGN KEY constraint failed' in str(e.orig):
            flash('Cannot delete payment type as it is linked to transactions.', 'danger')
        else:
            flash('An error occurred while deleting the payment type. Please try again.', 'danger')
    return redirect(url_for('payment_types'))

# --- User Account View ---
@app.route('/account')
@login_required
def account():
    acct = Account.query.filter_by(user_id=current_user.id).first()
    return render_template('account.html', account=acct)

# --- Admin Account Management ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or getattr(current_user, 'user_type', None) != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/accounts')
@login_required
@admin_required
def accounts():
    q = request.args.get('q', '')
    query = Account.query.join(User)
    if q:
        query = query.filter((User.username.ilike(f'%{q}%')) | (Account.account_number.ilike(f'%{q}%')))
    accounts = query.order_by(Account.id).all()
    return render_template('accounts.html', accounts=accounts, q=q)

@app.route('/accounts/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_account():
    users = User.query.all()
    if request.method == 'POST':
        user_id = int(request.form['user_id'])
        balance = float(request.form['balance'])
        if Account.query.filter_by(user_id=user_id).first():
            flash('User already has an account.', 'error')
        else:
            acct = Account(user_id=user_id, balance=balance)
            db.session.add(acct)
            try:
                db.session.commit()
                flash('Account added!', 'success')
                return redirect(url_for('accounts'))
            except IntegrityError as e:
                db.session.rollback()
                if 'UNIQUE constraint failed' in str(e.orig):
                    flash('User already has an account.', 'danger')
                else:
                    flash('An error occurred while adding the account. Please try again.', 'danger')
    return render_template('add_edit_account.html', users=users, title='Add Account')

@app.route('/accounts/edit/<int:account_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_account(account_id):
    acct = Account.query.get_or_404(account_id)
    users = User.query.all()
    if request.method == 'POST':
        acct.user_id = int(request.form['user_id'])
        # account_number is not editable
        acct.balance = float(request.form['balance'])
        try:
            db.session.commit()
            flash('Account updated!', 'success')
            return redirect(url_for('accounts'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('User already has an account.', 'danger')
            else:
                flash('An error occurred while updating the account. Please try again.', 'danger')
    return render_template('add_edit_account.html', account=acct, users=users, title='Edit Account')

@app.route('/accounts/delete/<int:account_id>', methods=['POST'])
@login_required
@admin_required
def delete_account(account_id):
    acct = Account.query.get_or_404(account_id)
    db.session.delete(acct)
    try:
        db.session.commit()
        flash('Account deleted!', 'success')
    except IntegrityError as e:
        db.session.rollback()
        if 'FOREIGN KEY constraint failed' in str(e.orig):
            flash('Cannot delete account as it is linked to transactions.', 'danger')
        else:
            flash('An error occurred while deleting the account. Please try again.', 'danger')
    return redirect(url_for('accounts'))

# --- Admin Transaction Management ---
@app.route('/transactions/all')
@login_required
@admin_required
def all_transactions():
    q = request.args.get('q', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    txn_type = request.args.get('txn_type', '')
    payment_type_id = request.args.get('payment_type_id', '')
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    status = request.args.get('status', '')

    txns = Transaction.query.join(Account).join(User)
    if q:
        txns = txns.filter(
            (User.username.ilike(f'%{q}%')) |
            (Transaction.reference_number.ilike(f'%{q}%')) |
            (Account.account_number.ilike(f'%{q}%'))
        )
    if start_date:
        from datetime import datetime
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            txns = txns.filter(Transaction.transaction_date >= start)
        except Exception:
            pass
    if end_date:
        from datetime import datetime, timedelta
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            txns = txns.filter(Transaction.transaction_date <= end)
        except Exception:
            pass
    if txn_type:
        txns = txns.filter(Transaction.transaction_type == txn_type)
    if payment_type_id:
        txns = txns.filter(Transaction.payment_type_id == int(payment_type_id))
    if min_amount:
        try:
            txns = txns.filter(Transaction.amount >= float(min_amount))
        except Exception:
            pass
    if max_amount:
        try:
            txns = txns.filter(Transaction.amount <= float(max_amount))
        except Exception:
            pass
    if status:
        txns = txns.filter(Transaction.status == status)
    txns = txns.order_by(Transaction.transaction_date.desc()).all()
    payment_types = PaymentType.query.filter_by(is_active=True).all()
    statuses = ['pending', 'completed', 'failed']
    return render_template('all_transactions.html', transactions=txns, q=q, start_date=start_date, end_date=end_date, txn_type=txn_type, payment_type_id=payment_type_id, min_amount=min_amount, max_amount=max_amount, status=status, payment_types=payment_types, statuses=statuses)

@app.route('/transactions/export')
@login_required
@admin_required
def export_transactions():
    q = request.args.get('q', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    txn_type = request.args.get('txn_type', '')
    payment_type_id = request.args.get('payment_type_id', '')
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    status = request.args.get('status', '')

    txns = Transaction.query.join(Account).join(User)
    if q:
        txns = txns.filter(
            (User.username.ilike(f'%{q}%')) |
            (Transaction.reference_number.ilike(f'%{q}%')) |
            (Account.account_number.ilike(f'%{q}%'))
        )
    if start_date:
        from datetime import datetime
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            txns = txns.filter(Transaction.transaction_date >= start)
        except Exception:
            pass
    if end_date:
        from datetime import datetime, timedelta
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            txns = txns.filter(Transaction.transaction_date <= end)
        except Exception:
            pass
    if txn_type:
        txns = txns.filter(Transaction.transaction_type == txn_type)
    if payment_type_id:
        txns = txns.filter(Transaction.payment_type_id == int(payment_type_id))
    if min_amount:
        try:
            txns = txns.filter(Transaction.amount >= float(min_amount))
        except Exception:
            pass
    if max_amount:
        try:
            txns = txns.filter(Transaction.amount <= float(max_amount))
        except Exception:
            pass
    if status:
        txns = txns.filter(Transaction.status == status)
    txns = txns.order_by(Transaction.transaction_date.desc()).all()

    # Prepare CSV
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(['User', 'Account Number', 'Date', 'Type', 'Amount', 'Reference', 'Payment Type', 'Balance After', 'Description', 'Status'])
    for txn in txns:
        cw.writerow([
            txn.account.user.username if txn.account and txn.account.user else '',
            txn.account.account_number if txn.account else '',
            txn.transaction_date.strftime('%Y-%m-%d %H:%M'),
            txn.transaction_type,
            txn.amount,
            txn.reference_number,
            txn.payment_type.name if txn.payment_type else '',
            txn.balance_after,
            txn.description or '',
            txn.status
        ])
    output = si.getvalue()
    si.close()
    return Response(
        output,
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment;filename=transactions.csv'
        }
    )

@app.route('/transactions/export_excel')
@login_required
@admin_required
def export_transactions_excel():
    q = request.args.get('q', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    txn_type = request.args.get('txn_type', '')
    payment_type_id = request.args.get('payment_type_id', '')
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    status = request.args.get('status', '')

    txns = Transaction.query.join(Account).join(User)
    if q:
        txns = txns.filter(
            (User.username.ilike(f'%{q}%')) |
            (Transaction.reference_number.ilike(f'%{q}%')) |
            (Account.account_number.ilike(f'%{q}%'))
        )
    if start_date:
        from datetime import datetime
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            txns = txns.filter(Transaction.transaction_date >= start)
        except Exception:
            pass
    if end_date:
        from datetime import datetime, timedelta
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            txns = txns.filter(Transaction.transaction_date <= end)
        except Exception:
            pass
    if txn_type:
        txns = txns.filter(Transaction.transaction_type == txn_type)
    if payment_type_id:
        txns = txns.filter(Transaction.payment_type_id == int(payment_type_id))
    if min_amount:
        try:
            txns = txns.filter(Transaction.amount >= float(min_amount))
        except Exception:
            pass
    if max_amount:
        try:
            txns = txns.filter(Transaction.amount <= float(max_amount))
        except Exception:
            pass
    if status:
        txns = txns.filter(Transaction.status == status)
    txns = txns.order_by(Transaction.transaction_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    headers = ['User', 'Account Number', 'Date', 'Type', 'Amount', 'Reference', 'Payment Type', 'Balance After', 'Description', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for txn in txns:
        ws.append([
            txn.account.user.username if txn.account and txn.account.user else '',
            txn.account.account_number if txn.account else '',
            txn.transaction_date.strftime('%Y-%m-%d %H:%M'),
            txn.transaction_type,
            txn.amount,
            txn.reference_number,
            txn.payment_type.name if txn.payment_type else '',
            txn.balance_after,
            txn.description or '',
            txn.status
        ])
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment;filename=transactions.xlsx'
        }
    )

@app.route('/transactions/export_pdf')
@login_required
@admin_required
def export_transactions_pdf():
    q = request.args.get('q', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    txn_type = request.args.get('txn_type', '')
    payment_type_id = request.args.get('payment_type_id', '')
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    status = request.args.get('status', '')

    txns = Transaction.query.join(Account).join(User)
    if q:
        txns = txns.filter(
            (User.username.ilike(f'%{q}%')) |
            (Transaction.reference_number.ilike(f'%{q}%')) |
            (Account.account_number.ilike(f'%{q}%'))
        )
    if start_date:
        from datetime import datetime
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            txns = txns.filter(Transaction.transaction_date >= start)
        except Exception:
            pass
    if end_date:
        from datetime import datetime, timedelta
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            txns = txns.filter(Transaction.transaction_date <= end)
        except Exception:
            pass
    if txn_type:
        txns = txns.filter(Transaction.transaction_type == txn_type)
    if payment_type_id:
        txns = txns.filter(Transaction.payment_type_id == int(payment_type_id))
    if min_amount:
        try:
            txns = txns.filter(Transaction.amount >= float(min_amount))
        except Exception:
            pass
    if max_amount:
        try:
            txns = txns.filter(Transaction.amount <= float(max_amount))
        except Exception:
            pass
    if status:
        txns = txns.filter(Transaction.status == status)
    txns = txns.order_by(Transaction.transaction_date.desc()).all()

    # Prepare PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    data = [[
        'User', 'Account Number', 'Date', 'Type', 'Amount', 'Reference', 'Payment Type', 'Balance After', 'Description', 'Status']
    ]
    for txn in txns:
        data.append([
            txn.account.user.username if txn.account and txn.account.user else '',
            txn.account.account_number if txn.account else '',
            txn.transaction_date.strftime('%Y-%m-%d %H:%M'),
            txn.transaction_type,
            txn.amount,
            txn.reference_number,
            txn.payment_type.name if txn.payment_type else '',
            txn.balance_after,
            txn.description or '',
            txn.status
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements = [table]
    doc.build(elements)
    output.seek(0)
    return Response(
        output,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': 'attachment;filename=transactions.pdf'
        }
    )

@app.route('/transactions/edit/<int:txn_id>', methods=['GET', 'POST'])
@login_required
def edit_transaction(txn_id):
    txn = Transaction.query.get_or_404(txn_id)
    form = TransactionForm(obj=txn)
    form.payment_type.choices = [(pt.id, pt.name) for pt in PaymentType.query.filter_by(is_active=True).all()]
    if form.validate_on_submit():
        txn.amount = form.amount.data
        txn.transaction_type = form.transaction_type.data
        txn.reference_number = form.reference_number.data
        txn.payment_type_id = form.payment_type.data
        txn.description = form.description.data
        # Recalculate balance_after (for admin, you may want to recalc all after editing)
        account = txn.account
        if txn.transaction_type == 'credit':
            txn.balance_after = account.balance + txn.amount
        else:
            txn.balance_after = account.balance - txn.amount
        try:
            db.session.commit()
            flash('Transaction updated!', 'success')
            return redirect(url_for('all_transactions'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('Reference number already exists. Please use a unique one.', 'danger')
            else:
                flash('An error occurred while updating the transaction. Please try again.', 'danger')
    return render_template('add_transaction.html', form=form, edit=True)

@app.route('/transactions/delete/<int:txn_id>', methods=['POST'])
@login_required
def delete_transaction(txn_id):
    txn = Transaction.query.get_or_404(txn_id)
    db.session.delete(txn)
    try:
        db.session.commit()
        flash('Transaction deleted!', 'success')
    except IntegrityError as e:
        db.session.rollback()
        if 'FOREIGN KEY constraint failed' in str(e.orig):
            flash('Cannot delete transaction as it is linked to an account.', 'danger')
        else:
            flash('An error occurred while deleting the transaction. Please try again.', 'danger')
    return redirect(url_for('all_transactions'))

@app.route('/package_items')
@login_required
def package_items():
    items = db.session.query(PackageItem).join(Package).order_by(Package.name, PackageItem.price).all()
    return render_template('package_items.html', items=items)

@app.route('/package_items/add', methods=['GET', 'POST'])
@login_required
def add_package_item():
    form = PackageItemForm()
    form.package_id.choices = [(p.id, p.name) for p in Package.query.filter_by(is_active=True).all()]
    if form.validate_on_submit():
        item = PackageItem(
            package_id=form.package_id.data,
            description=form.description.data,
            price=form.price.data,
            is_active=form.is_active.data
        )
        db.session.add(item)
        try:
            db.session.commit()
            flash('Package item added!', 'success')
            return redirect(url_for('package_items'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A package item with that description already exists for this package.', 'danger')
            else:
                flash('An error occurred while adding the package item. Please try again.', 'danger')
    return render_template('add_edit_package_item.html', form=form, title='Add Package Item')

@app.route('/package_items/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_package_item(item_id):
    item = PackageItem.query.get_or_404(item_id)
    form = PackageItemForm(obj=item)
    form.package_id.choices = [(p.id, p.name) for p in Package.query.filter_by(is_active=True).all()]
    if form.validate_on_submit():
        item.package_id = form.package_id.data
        item.description = form.description.data
        item.price = form.price.data
        item.is_active = form.is_active.data
        try:
            db.session.commit()
            flash('Package item updated!', 'success')
            return redirect(url_for('package_items'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('A package item with that description already exists for this package.', 'danger')
            else:
                flash('An error occurred while updating the package item. Please try again.', 'danger')
    return render_template('add_edit_package_item.html', form=form, title='Edit Package Item')

@app.route('/package_items/delete/<int:item_id>', methods=['POST'])
@login_required
def delete_package_item(item_id):
    item = PackageItem.query.get_or_404(item_id)
    db.session.delete(item)
    try:
        db.session.commit()
        flash('Package item deleted!', 'success')
    except IntegrityError as e:
        db.session.rollback()
        if 'FOREIGN KEY constraint failed' in str(e.orig):
            flash('Cannot delete package item as it is linked to a transaction.', 'danger')
        else:
            flash('An error occurred while deleting the package item. Please try again.', 'danger')
    return redirect(url_for('package_items'))

@app.route('/api/package_items/<int:package_id>')
@login_required
def api_package_items(package_id):
    items = PackageItem.query.filter_by(package_id=package_id, is_active=True).all()
    return jsonify({
        'items': [
            {'id': item.id, 'label': f"{item.description} - ${item.price}"} for item in items
        ]
    })

@app.route('/purchase_package', methods=['GET', 'POST'])
@login_required
def purchase_package():
    packages = Package.query.filter_by(is_active=True).all()
    package_id = request.args.get('package_id')
    form = PackageItemForm()
    form.package_id.choices = [(p.id, p.name) for p in packages]
    
    if package_id:
        form.package_id.data = int(package_id)  # Convert to integer for SelectField coerce=int
    
    topup_for = request.form.get('topup_for', '')
    if request.method == 'POST':
        package_id = request.form.get('package_id')
        package_item_id = request.form.get('package_item_id')
        topup_for = request.form.get('topup_for', '')
        # Phone number validation: allow +, 10-15 digits 
        phone_pattern = re.compile(r'^\+?\d{10,15}$')
        if not phone_pattern.match(topup_for):
            flash('Please enter a valid phone number (10-15 digits, numbers only, optional + at start).', 'danger')
            return render_template('purchase_package.html', packages=packages, topup_for=topup_for)
        try:
            package = Package.query.get(package_id)
            package_item = PackageItem.query.get(package_item_id) if package_item_id else None
            if not package or (package_item_id and not package_item):
                flash('Invalid package or package item selected.', 'danger')
                return redirect(url_for('purchase_package'))
            amount = package_item.price if package_item else 0
            account = Account.query.filter_by(user_id=current_user.id).first()
            if not account:
                flash('Account not found.', 'danger')
                return redirect(url_for('purchase_package'))
            if account.balance < amount:
                flash('Insufficient balance.', 'danger')
                return redirect(url_for('purchase_package'))
            account.balance -= amount
            order = Orders(
                account_id=account.id,
                package_id=package.id,
                package_item_id=package_item.id if package_item else None,
                amount=amount,
                status='completed',
                topup_for=topup_for
            )
            db.session.add(order)
            db.session.commit()
            flash('Purchase successful!', 'success')
            return redirect(url_for('orders'))
        except IntegrityError as e:
            db.session.rollback()
            flash('A database error occurred. Please try again.', 'danger')
            return redirect(url_for('purchase_package'))
        except Exception as e:
            db.session.rollback()
            flash(f'An unexpected error occurred: {str(e)}', 'danger')
            return redirect(url_for('purchase_package'))
    return render_template('purchase_package.html', form=form, packages=packages, topup_for=topup_for)

@app.route('/orders')
@login_required
def orders():
    account = Account.query.filter_by(user_id=current_user.id).first()
    if not account:
        flash('No account found for this user.', 'error')
        return redirect(url_for('home'))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status', '')
    topup_for = request.args.get('topup_for', '')
    sort = request.args.get('sort', 'created_at')
    direction = request.args.get('direction', 'desc')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    query = Orders.query.filter_by(account_id=account.id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Orders.created_at >= start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            query = query.filter(Orders.created_at <= end)
        except ValueError:
            pass
    if status:
        query = query.filter(Orders.status == status)
    if topup_for:
        # Normalize phone number for filtering (remove spaces, dashes, etc.)
        norm_topup_for = re.sub(r'[^\d+]', '', topup_for)
        query = query.filter(Orders.topup_for == norm_topup_for)
    sort_column = getattr(Orders, sort, Orders.created_at)
    if direction == 'asc':
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    orders = pagination.items
    return render_template('orders.html', orders=orders, pagination=pagination, start_date=start_date, end_date=end_date, status=status, topup_for=topup_for, sort=sort, direction=direction, per_page=per_page)

@app.route('/admin/orders')
@login_required
@admin_required
def admin_orders():
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    sort = request.args.get('sort', 'created_at')
    direction = request.args.get('direction', 'desc')
    query = Orders.query
    if user_id:
        query = query.join(Account).filter(Account.user_id == user_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Orders.created_at >= start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            query = query.filter(Orders.created_at <= end)
        except ValueError:
            pass
    # Sorting
    sort_column = getattr(Orders, sort, Orders.created_at)
    if direction == 'asc':
        query = query.order_by(asc(sort_column))
    else:
        query = query.order_by(desc(sort_column))
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    orders = pagination.items
    from models.user import User
    users = User.query.order_by(User.username).all()
    return render_template('admin_orders.html', orders=orders, users=users, user_id=user_id, start_date=start_date, end_date=end_date, pagination=pagination, sort=sort, direction=direction, per_page=per_page)

@app.route('/admin/orders/update_status', methods=['POST'])
def admin_update_order_status():
    order_id = request.form.get('order_id')
    new_status = request.form.get('status')
    order = Orders.query.get(order_id)
    if order and new_status in ['Pending', 'Processing', 'Completed']:
        order.status = new_status
        db.session.commit()
        flash('Order status updated.', 'success')
    else:
        flash('Invalid order or status.', 'danger')
    return redirect(request.referrer or url_for('admin_orders'))

@app.route('/admin/orders/<int:order_id>/json')
@login_required
@admin_required
def admin_order_json(order_id):
    order = Orders.query.get_or_404(order_id)
    data = {
        'id': order.id,
        'date': order.created_at.strftime('%Y-%m-%d %H:%M'),
        'user': order.account.user.username if order.account and order.account.user else '',
        'package': order.package.name,
        'item': order.package_item.description if order.package_item else '',
        'amount': order.amount,
        'status': order.status,
        'account_number': order.account.account_number if order.account else '',
        'user_email': order.account.user.email if order.account and order.account.user else '',
        'user_phone': order.account.user.phone if order.account and order.account.user and hasattr(order.account.user, 'phone') else '',
        'description': order.package_item.description if order.package_item else '',
        'topup_for': order.topup_for
    }
    return jsonify(data)

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    from forms.registerform import EditUserForm
    form = EditUserForm(obj=current_user)
    # Hide user_type field for normal users
    if current_user.user_type != 'admin':
        form.user_type.render_kw = {'disabled': True}
    if form.validate_on_submit():
        current_user.first_name = form.first_name.data
        current_user.last_name = form.last_name.data
        current_user.username = form.username.data
        current_user.email = form.email.data
        current_user.phone_number = form.phone_number.data
        # Only admin can change user_type
        if current_user.user_type == 'admin':
            current_user.user_type = form.user_type.data
        if form.password.data:
            current_user.set_password(form.password.data)
        try:
            db.session.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('profile'))
        except IntegrityError as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e.orig):
                flash('Username or email is already in use. Please choose a different one.', 'danger')
            else:
                flash('An error occurred while updating the profile. Please try again.', 'danger')
    return render_template('edit_profile.html', form=form)

@app.route('/profile/delete', methods=['POST'])
@login_required
def delete_profile():
    user = current_user
    logout_user()
    try:
        db.session.delete(user)
        db.session.commit()
        flash('Your account has been deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting your account. Please try again.', 'error')
    return redirect(url_for('home'))

def get_filtered_orders(user_id=None, start_date=None, end_date=None):
    query = Orders.query
    if user_id:
        query = query.join(Account).filter(Account.user_id == user_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Orders.created_at >= start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            query = query.filter(Orders.created_at <= end)
        except ValueError:
            pass
    return query.order_by(Orders.created_at.desc()).all()

@app.route('/admin/orders/export/excel')
@login_required
@admin_required
def export_orders_excel():
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    orders = get_filtered_orders(user_id, start_date, end_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'
    ws.append(['Date', 'User', 'Package', 'Item', 'Amount', 'Status'])
    for order in orders:
        ws.append([
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.account.user.username if order.account and order.account.user else '',
            order.package.name,
            order.package_item.description if order.package_item else '',
            order.amount,
            order.status
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='orders.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/orders/export/pdf')
@login_required
@admin_required
def export_orders_pdf():
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    orders = get_filtered_orders(user_id, start_date, end_date)
    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter
    y = height - 40
    p.setFont('Helvetica-Bold', 14)
    p.drawString(40, y, 'Orders')
    y -= 30
    p.setFont('Helvetica-Bold', 10)
    headers = ['Date', 'User', 'Package', 'Item', 'Amount', 'Status']
    for i, header in enumerate(headers):
        p.drawString(40 + i*90, y, header)
    y -= 20
    p.setFont('Helvetica', 10)
    for order in orders:
        if y < 50:
            p.showPage()
            y = height - 40
        row = [
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.account.user.username if order.account and order.account.user else '',
            order.package.name,
            order.package_item.description if order.package_item else '',
            str(order.amount),
            order.status
        ]
        for i, value in enumerate(row):
            p.drawString(40 + i*90, y, value)
        y -= 18
    p.save()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='orders.pdf', mimetype='application/pdf')

if __name__ == "__main__":
    app.run(debug=True)